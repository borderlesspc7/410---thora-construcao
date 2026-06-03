"""
Geração de XLSX com abas condicionais conforme modelosSelecionados.
"""

from __future__ import annotations

import re
import uuid
from pathlib import Path
from typing import Any, Dict, List, Tuple

from services.analitico_normalize import normalize_hierarchical_analitico
from services.openai_service import _coerce_bdi, _coerce_number

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore

DEFAULT_MODELS = {
    "analitico": False,
    "sintetico": False,
    "curva_abc": True,
}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
GROUP_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
ANALITICO_GROUP_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
GROUP_FONT = Font(bold=True, size=10)
COMP_FONT = Font(italic=True, size=9, color="475569")
TOTAL_FILL = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
ZEBRA_LIGHT = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
ZEBRA_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
CLASS_FILLS = {
    "A": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "B": PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid"),
    "C": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
}
CLASS_FONTS = {
    "A": Font(bold=True, color="991B1B"),
    "B": Font(bold=True, color="854D0E"),
    "C": Font(bold=True, color="065F46"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def normalize_models_selection(raw: Dict[str, bool] | None) -> Dict[str, bool]:
    if not raw:
        return dict(DEFAULT_MODELS)
    return {
        "analitico": bool(raw.get("analitico", raw.get("analítico", True))),
        "sintetico": bool(raw.get("sintetico", raw.get("sintético", True))),
        "curva_abc": bool(raw.get("curva_abc", raw.get("abc", True))),
    }


def _bdi_factor(bdi_percent: float) -> float:
    return 1.0 + (bdi_percent / 100.0) if bdi_percent > 0 else 1.0


def _resolve_tipo_linha(raw: Dict[str, Any]) -> str:
    tipo = str(raw.get("tipo_linha") or raw.get("tipo") or "item").strip().lower()
    if tipo in ("grupo", "titulo", "título", "title"):
        return "grupo"
    if tipo in ("composicao", "composição", "insumo", "subitem"):
        return "composicao"
    return "item"


def _is_group_row(raw: Dict[str, Any]) -> bool:
    tipo = _resolve_tipo_linha(raw)
    desc = str(raw.get("description") or raw.get("descricao") or "").strip().lower()
    return tipo == "grupo" or "total do grupo" in desc


def _is_analitico_grupo_visual(row_data: Dict[str, Any]) -> bool:
    """Grupo no Excel: sem Código e sem Unidade (cinza + negrito)."""
    if _is_group_row(row_data):
        return True
    codigo = str(row_data.get("code") or row_data.get("codigo") or "").strip()
    unidade = str(row_data.get("unit") or row_data.get("unidade") or "").strip()
    return not codigo and not unidade


def _is_composicao_row(raw: Dict[str, Any]) -> bool:
    return _resolve_tipo_linha(raw) == "composicao"


def _is_executive_row(raw: Dict[str, Any]) -> bool:
    return _resolve_tipo_linha(raw) == "item" and not _is_group_row(raw)


def prepare_hierarchical_analitico_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Preserva ordem, tipagem, totais e numeração hierárquica (Python, não IA)."""
    payload: List[Dict[str, Any]] = []
    for idx, raw in enumerate(items):
        if not isinstance(raw, dict):
            continue
        payload.append(
            {
                **raw,
                "item_numero": str(
                    raw.get("item_numero") or raw.get("item") or raw.get("id") or ""
                ).strip(),
                "descricao": str(raw.get("descricao") or raw.get("description") or "").strip(),
                "codigo": str(raw.get("codigo") or raw.get("code") or "").strip(),
                "unidade": str(raw.get("unidade") or raw.get("unit") or "").strip(),
                "quantidade": _coerce_number(raw.get("quantidade") or raw.get("qty")),
                "valor_unitario": _coerce_number(
                    raw.get("valor_unitario")
                    or raw.get("unitPrice")
                    or raw.get("unit_com_bdi")
                ),
                "valor_total": _coerce_number(
                    raw.get("valor_total") or raw.get("total_com_bdi") or raw.get("totalValue")
                ),
                "bdi": _coerce_bdi(raw.get("bdi") or raw.get("BDI")),
                "tipo_linha": raw.get("tipo_linha") or raw.get("tipo"),
                "rotulo_linha": str(raw.get("rotulo_linha") or "").strip(),
                "tipo_categoria": str(raw.get("tipo_categoria") or "").strip(),
                "porcentagem": _coerce_number(raw.get("porcentagem") or raw.get("percentual")),
                "banco": str(raw.get("banco") or "").strip(),
                "_order": idx,
            }
        )

    normalized = normalize_hierarchical_analitico(payload)
    rows: List[Dict[str, Any]] = []
    for row in normalized:
        base = _normalize_base_row(row)
        base["item_numero"] = str(row.get("item_numero") or row.get("item") or "").strip()
        base["rotulo_linha"] = str(row.get("rotulo_linha") or "").strip()
        base["tipo_categoria"] = str(row.get("tipo_categoria") or "").strip()
        base["porcentagem"] = _coerce_number(row.get("porcentagem") or 0)
        base["tipo_linha"] = str(row.get("tipo_linha") or "item")
        base["banco"] = str(row.get("banco") or "").strip()
        base["qty"] = _coerce_number(row.get("quantidade") or row.get("qty"))
        base["unit"] = str(row.get("unidade") or row.get("unit") or "").strip()
        base["unit_com_bdi"] = _coerce_number(
            row.get("valor_unitario") or row.get("unit_com_bdi")
        )
        base["total_com_bdi"] = _coerce_number(
            row.get("valor_total") or row.get("total_com_bdi")
        )
        base["code"] = str(row.get("codigo") or row.get("code") or "").strip()
        base["description"] = str(row.get("descricao") or row.get("description") or "").strip()
        base["_order"] = row.get("_order")
        rows.append(base)
    return rows


def _line_total_com_bdi(raw: Dict[str, Any]) -> float:
    bdi = _coerce_bdi(raw.get("bdi") or raw.get("BDI"))
    qty = _coerce_number(raw.get("qty") or raw.get("quantidade") or raw.get("quantity"))
    unit_com_bdi = _coerce_number(
        raw.get("unitPrice") or raw.get("valor_unitario") or raw.get("unitValue")
    )
    total_com_bdi = _coerce_number(
        raw.get("lineTotal")
        or raw.get("line_total")
        or raw.get("totalValue")
        or raw.get("valor_total")
    )
    if total_com_bdi <= 0 and qty > 0 and unit_com_bdi > 0:
        total_com_bdi = qty * unit_com_bdi
    if total_com_bdi <= 0 and qty > 0:
        unit_sem = unit_com_bdi / _bdi_factor(bdi) if _bdi_factor(bdi) > 0 else unit_com_bdi
        total_com_bdi = qty * unit_com_bdi if unit_com_bdi > 0 else qty * unit_sem * _bdi_factor(bdi)
    return total_com_bdi


def _normalize_base_row(raw: Dict[str, Any]) -> Dict[str, Any]:
    bdi = _coerce_bdi(raw.get("bdi") or raw.get("BDI"))
    qty = _coerce_number(raw.get("qty") or raw.get("quantidade") or raw.get("quantity"))
    unit_com_bdi = _coerce_number(
        raw.get("unitPrice") or raw.get("valor_unitario") or raw.get("unitValue")
    )
    total_com_bdi = _line_total_com_bdi(raw)
    factor = _bdi_factor(bdi)
    unit_sem_bdi = unit_com_bdi / factor if factor > 0 else unit_com_bdi
    if unit_com_bdi <= 0 and qty > 0 and total_com_bdi > 0:
        unit_com_bdi = total_com_bdi / qty
    if unit_sem_bdi <= 0 and qty > 0 and total_com_bdi > 0:
        unit_sem_bdi = (total_com_bdi / factor) / qty if factor > 0 else total_com_bdi / qty

    return {
        "code": str(raw.get("code") or raw.get("codigo") or "").strip(),
        "description": str(raw.get("description") or raw.get("descricao") or "").strip(),
        "bdi": bdi,
        "unit": str(raw.get("unit") or raw.get("unidade") or "").strip(),
        "qty": qty,
        "unit_com_bdi": unit_com_bdi,
        "total_com_bdi": total_com_bdi,
        "grupo": str(raw.get("grupo") or "").strip(),
        "classification": str(raw.get("classification") or raw.get("class") or "")
        .strip()
        .upper(),
        "accumulated_percentage": raw.get("accumulated_percentage"),
        "_order": raw.get("_order"),
    }


def prepare_analitico_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Ordem original (campo item/id); apenas itens executivos; sem ABC."""
    rows: List[Dict[str, Any]] = []
    for idx, raw in enumerate(items):
        if not isinstance(raw, dict) or not _is_executive_row(raw):
            continue
        row = _normalize_base_row(raw)
        row["_order"] = _coerce_number(raw.get("item") or raw.get("id") or idx)
        rows.append(row)
    rows.sort(key=lambda r: (r["_order"], r["code"]))
    return rows


def prepare_curva_abc_rows(items: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], float]:
    """Ordenação por valor decrescente + percentuais e classificação A/B/C."""
    prepared: List[Dict[str, Any]] = []
    for raw in items:
        if not isinstance(raw, dict) or not _is_executive_row(raw):
            continue
        prepared.append(_normalize_base_row(raw))

    prepared.sort(key=lambda row: row["total_com_bdi"], reverse=True)
    total_geral = sum(row["total_com_bdi"] for row in prepared)

    accumulated = 0.0
    for row in prepared:
        percent = (row["total_com_bdi"] / total_geral * 100.0) if total_geral > 0 else 0.0
        pct_before = accumulated
        accumulated += percent
        row["percent"] = percent
        acc_front = row.get("accumulated_percentage")
        row["accumulated"] = (
            float(acc_front) if acc_front is not None and acc_front != "" else accumulated
        )
        if not row["classification"]:
            if pct_before < 80:
                row["classification"] = "A"
            elif pct_before < 95:
                row["classification"] = "B"
            else:
                row["classification"] = "C"

    return prepared, total_geral


def prepare_sintetico_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Linhas de grupo/título com totais; sem itens executivos filhos."""
    sintetico: List[Dict[str, Any]] = []

    for raw in items:
        if not isinstance(raw, dict) or not _is_group_row(raw):
            continue
        row = _normalize_base_row(raw)
        if row["total_com_bdi"] <= 0:
            grupo_key = row["description"] or row["code"]
            row["total_com_bdi"] = sum(
                _line_total_com_bdi(r)
                for r in items
                if isinstance(r, dict)
                and _is_executive_row(r)
                and str(r.get("grupo") or "").strip() == grupo_key
            )
        sintetico.append(row)

    if sintetico:
        return sintetico

    totals_by_grupo: Dict[str, float] = {}
    labels_by_grupo: Dict[str, str] = {}
    for raw in items:
        if not isinstance(raw, dict) or not _is_executive_row(raw):
            continue
        key = str(raw.get("grupo") or "Geral").strip() or "Geral"
        totals_by_grupo[key] = totals_by_grupo.get(key, 0.0) + _line_total_com_bdi(raw)
        labels_by_grupo[key] = key

    for key, total in totals_by_grupo.items():
        sintetico.append(
            {
                "code": "",
                "description": labels_by_grupo.get(key, key),
                "bdi": 0.0,
                "unit": "",
                "qty": 0.0,
                "unit_com_bdi": 0.0,
                "total_com_bdi": total,
                "grupo": key,
            }
        )

    return sintetico


def _write_header_row(ws, headers: List[str]) -> None:
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def _apply_col_widths(ws, widths: Dict[str, float]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _write_novacap_metadata_header(
    ws,
    *,
    nome_obra: str | None = None,
    bancos_referencia: str | None = None,
    bdi_percent: float | None = None,
    encargos_sociais: str | None = None,
) -> int:
    """Cabeçalho estilo NOVACAP (linhas 1-3). Retorna linha inicial dos dados."""
    ws.cell(row=1, column=4).value = "Obra"
    ws.cell(row=1, column=5).value = "Bancos"
    ws.cell(row=1, column=7).value = "B.D.I."
    ws.cell(row=1, column=9).value = "Encargos Sociais"
    for col in (4, 5, 7, 9):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, size=10)

    ws.cell(row=2, column=4).value = nome_obra or "—"
    ws.cell(row=2, column=5).value = bancos_referencia or "SINAPI / SICRO"
    ws.cell(row=2, column=7).value = (
        f"{bdi_percent:.2f}%".replace(".", ",") if bdi_percent is not None else "—"
    )
    ws.cell(row=2, column=9).value = encargos_sociais or "—"
    for col in (4, 5, 7, 9):
        ws.cell(row=2, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    title_cell = ws.cell(row=3, column=1)
    title_cell.value = "Planilha Orçamentária Analítica"
    title_cell.font = Font(bold=True, size=12)
    return 4


def gerar_aba_analitica(
    ws,
    rows: List[Dict[str, Any]],
    *,
    nome_obra: str | None = None,
    bancos_referencia: str | None = None,
    bdi_percent: float | None = None,
    encargos_sociais: str | None = None,
) -> None:
    """
    Planilha analítica — colunas do edital (A–H):
    Item | Código | Banco | Descrição | Und | Quant. | Valor Unit | Total
    """
    if not bancos_referencia:
        bancos = sorted(
            {
                str(r.get("banco") or "").strip()
                for r in rows
                if str(r.get("banco") or "").strip()
            }
        )
        bancos_referencia = "\n".join(bancos) if bancos else "SINAPI / SICRO"

    if bdi_percent is None:
        bdi_values = [float(r.get("bdi") or 0) for r in rows if float(r.get("bdi") or 0) > 0]
        bdi_percent = sum(bdi_values) / len(bdi_values) if bdi_values else None

    meta_end = _write_novacap_metadata_header(
        ws,
        nome_obra=nome_obra,
        bancos_referencia=bancos_referencia,
        bdi_percent=bdi_percent,
        encargos_sociais=encargos_sociais,
    )

    header_row = meta_end
    headers = [
        "Item",
        "Código",
        "Tipo",
        "Banco",
        "Descrição",
        "Und",
        "Quant.",
        "Valor Unit",
        "Total",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    data_start = header_row + 1
    open_grupos: List[Tuple[int, int]] = []

    def _close_grupo(grupo_row: int, child_start: int, child_end: int) -> None:
        if child_end < child_start:
            return
        total_col = get_column_letter(9)
        total_cell = ws.cell(row=grupo_row, column=9)
        total_cell.value = f"=SUM({total_col}{child_start}:{total_col}{child_end})"
        total_cell.number_format = "#,##0.00"
        total_cell.font = GROUP_FONT
        total_cell.alignment = Alignment(horizontal="right", vertical="center")

    for idx, row_data in enumerate(rows):
        row_num = data_start + idx
        tipo = str(row_data.get("tipo_linha") or "item").lower()
        is_grupo = _is_analitico_grupo_visual(row_data)
        is_comp = tipo == "composicao" and not is_grupo

        item_num = str(row_data.get("item_numero") or "").strip()
        rotulo = str(row_data.get("rotulo_linha") or "").strip()
        col_a = item_num or rotulo
        codigo = str(row_data.get("code") or row_data.get("codigo") or "").strip()
        tipo_label = "Grupo" if is_grupo else ("Composição" if is_comp else "Item")
        banco = str(row_data.get("banco") or "").strip()
        descricao = str(row_data.get("description") or row_data.get("descricao") or "").strip()
        unidade = str(row_data.get("unit") or row_data.get("unidade") or "").strip()
        qty = _coerce_number(row_data.get("qty") or row_data.get("quantidade"))
        bdi = _coerce_bdi(row_data.get("bdi") or row_data.get("BDI"))
        unit_val = _coerce_number(
            row_data.get("unitPrice")
            or row_data.get("valor_unitario")
            or row_data.get("unit_com_bdi")
        )

        row_fill = ANALITICO_GROUP_FILL if is_grupo else (ZEBRA_LIGHT if idx % 2 == 0 else ZEBRA_WHITE)

        if is_grupo:
            if open_grupos:
                prev_row, child_start = open_grupos.pop()
                _close_grupo(prev_row, child_start, row_num - 1)
            for col in range(1, 10):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = ANALITICO_GROUP_FILL
                cell.font = GROUP_FONT
                cell.border = THIN_BORDER
            ws.cell(row=row_num, column=1).value = col_a
            ws.cell(row=row_num, column=5).value = descricao
            total_cell = ws.cell(row=row_num, column=9)
            total_cell.font = GROUP_FONT
            total_cell.fill = ANALITICO_GROUP_FILL
            total_cell.border = THIN_BORDER
            total_cell.alignment = Alignment(horizontal="right", vertical="center")
            open_grupos.append((row_num, row_num + 1))
            continue

        values = {
            1: col_a,
            2: codigo,
            3: tipo_label,
            4: banco,
            5: descricao,
            6: unidade,
            7: qty if qty else None,
            8: unit_val if unit_val else None,
        }

        for col_num, value in values.items():
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = THIN_BORDER
            cell.fill = row_fill
            cell.value = value
            if col_num == 7:
                cell.number_format = "#,##0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_num == 8:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_num == 5 and is_comp:
                cell.font = COMP_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            elif col_num == 5:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col_num == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        total_cell = ws.cell(row=row_num, column=9)
        total_cell.border = THIN_BORDER
        total_cell.fill = row_fill
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.number_format = "#,##0.00"
        qty_col = get_column_letter(7)
        unit_col = get_column_letter(8)
        static_total = _coerce_number(
            row_data.get("total_com_bdi") or row_data.get("valor_total") or row_data.get("totalValue")
        )
        if qty > 0 and unit_val > 0:
            computed_total = round(qty * unit_val, 2)
            total_cell.value = computed_total if computed_total > 0 else static_total
        elif static_total > 0:
            total_cell.value = static_total

    last_row = data_start + len(rows) - 1
    while open_grupos:
        grupo_row, child_start = open_grupos.pop()
        _close_grupo(grupo_row, child_start, last_row)

    ws.freeze_panes = f"A{data_start}"
    _apply_col_widths(
        ws,
        {"A": 14, "B": 14, "C": 12, "D": 12, "E": 48, "F": 8, "G": 12, "H": 14, "I": 16},
    )


def _fill_analitico_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    gerar_aba_analitica(ws, rows)


def _fill_curva_abc_sheet(ws, rows: List[Dict[str, Any]], total_geral: float) -> None:
    headers = [
        "Código",
        "Descrição",
        "BDI (%)",
        "Unidade",
        "Quantidade",
        "Valor Unitário C/BDI",
        "Valor Total C/BDI",
        "%",
        "Acumulado",
        "Class.",
    ]
    _write_header_row(ws, headers)
    right_cols = {3, 5, 6, 7, 8, 9}
    center_cols = {4, 10}

    for idx, row_data in enumerate(rows):
        row_num = idx + 2
        stripe = ZEBRA_LIGHT if idx % 2 == 0 else ZEBRA_WHITE
        values = [
            row_data["code"],
            row_data["description"],
            row_data["bdi"],
            row_data["unit"],
            row_data["qty"],
            row_data["unit_com_bdi"],
            row_data["total_com_bdi"],
            row_data.get("percent", 0),
            row_data.get("accumulated", 0),
            row_data.get("classification", ""),
        ]
        formats = [
            None,
            None,
            '0.00"%"',
            None,
            "#,##0.0000",
            "#,##0.000",
            "#,##0.00",
            '0.00"%"',
            '0.00"%"',
            None,
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = THIN_BORDER
            if formats[col_num - 1]:
                cell.number_format = formats[col_num - 1]
            if col_num == 10:
                cls = str(row_data.get("classification") or "").strip().upper()
                cell.fill = CLASS_FILLS.get(cls, stripe)
                cell.font = CLASS_FONTS.get(cls, Font(bold=True))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = stripe
                if col_num in right_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_num in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    total_row = len(rows) + 3
    ws.cell(row=total_row, column=6).value = "TOTAL GERAL:"
    ws.cell(row=total_row, column=6).font = TOTAL_FONT
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=7).value = total_geral
    ws.cell(row=total_row, column=7).number_format = "#,##0.00"
    ws.cell(row=total_row, column=7).fill = TOTAL_FILL
    ws.cell(row=total_row, column=7).font = TOTAL_FONT
    ws.cell(row=total_row, column=7).border = THIN_BORDER

    _apply_col_widths(
        ws,
        {
            "A": 14,
            "B": 48,
            "C": 10,
            "D": 10,
            "E": 12,
            "F": 16,
            "G": 16,
            "H": 10,
            "I": 12,
            "J": 8,
        },
    )


def _fill_sintetico_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    headers = ["Código", "Grupo / Etapa", "Valor Total C/BDI"]
    _write_header_row(ws, headers)
    total_geral = 0.0

    for idx, row_data in enumerate(rows):
        row_num = idx + 2
        stripe = ZEBRA_LIGHT if idx % 2 == 0 else ZEBRA_WHITE
        values = [
            row_data["code"],
            row_data["description"],
            row_data["total_com_bdi"],
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = THIN_BORDER
            cell.fill = stripe
            if col_num == 3:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        total_geral += row_data["total_com_bdi"]

    total_row = len(rows) + 3
    ws.cell(row=total_row, column=2).value = "TOTAL GERAL:"
    ws.cell(row=total_row, column=2).font = TOTAL_FONT
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=3).value = total_geral
    ws.cell(row=total_row, column=3).number_format = "#,##0.00"
    ws.cell(row=total_row, column=3).fill = TOTAL_FILL
    ws.cell(row=total_row, column=3).font = TOTAL_FONT
    ws.cell(row=total_row, column=3).border = THIN_BORDER

    _apply_col_widths(ws, {"A": 14, "B": 56, "C": 20})


def build_export_workbook(
    items: List[Dict[str, Any]],
    modelos_selecionados: Dict[str, bool] | None,
    *,
    nome_projeto: str | None = None,
) -> Tuple[Any, List[str]]:
    if not Workbook:
        raise RuntimeError("openpyxl não está instalado")

    models = normalize_models_selection(modelos_selecionados)
    if not any(models.values()):
        models = dict(DEFAULT_MODELS)

    wb = Workbook()
    wb.remove(wb.active)
    sheets_created: List[str] = []

    if models.get("analitico"):
        analitico_rows = prepare_hierarchical_analitico_rows(items)
        if analitico_rows:
            ws = wb.create_sheet("Orçamento Analítico")
            gerar_aba_analitica(ws, analitico_rows, nome_obra=nome_projeto)
            sheets_created.append("Orçamento Analítico")

    if models.get("curva_abc"):
        abc_rows, total_geral = prepare_curva_abc_rows(items)
        if abc_rows:
            ws = wb.create_sheet("Curva ABC")
            _fill_curva_abc_sheet(ws, abc_rows, total_geral)
            sheets_created.append("Curva ABC")

    if models.get("sintetico"):
        sintetico_rows = prepare_sintetico_rows(items)
        if sintetico_rows:
            ws = wb.create_sheet("Orçamento Sintético")
            _fill_sintetico_sheet(ws, sintetico_rows)
            sheets_created.append("Orçamento Sintético")

    if not sheets_created:
        raise ValueError(
            "Nenhuma aba pôde ser gerada. Verifique os itens e os modelos selecionados."
        )

    return wb, sheets_created


def save_export_workbook(
    items: List[Dict[str, Any]],
    modelos_selecionados: Dict[str, bool] | None,
    temp_folder: Path,
    nome_projeto: str | None = None,
) -> Tuple[Path, str]:
    wb, _ = build_export_workbook(items, modelos_selecionados, nome_projeto=nome_projeto)
    stem = "orcamento"
    if nome_projeto and nome_projeto.strip():
        safe = re.sub(r"[^\w\s-]", "", nome_projeto.strip(), flags=re.UNICODE)
        safe = re.sub(r"\s+", "_", safe)[:40]
        if safe:
            stem = safe
    filename = f"{stem}_{uuid.uuid4().hex[:8]}.xlsx"
    file_path = temp_folder / filename
    wb.save(file_path)
    return file_path, filename
